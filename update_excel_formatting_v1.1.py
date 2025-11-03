import openpyxl
from openpyxl.styles import PatternFill, Color
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation

# --- Configuration Variables ---
INPUT_FILENAME = "activity_tracker.xlsx"
OUTPUT_FILENAME = "activity_tracker_formatted.xlsx"
STATUS_OPTIONS = ["COMPLETE", "INCOMPLETE", "WIP"]

# --- Color Definitions ---
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED_ACCENT_FILL = PatternFill(bgColor=Color(theme=5, tint=0.4), fill_type="solid")
OLIVE_ACCENT_FILL = PatternFill(bgColor=Color(theme=9, tint=0.4), fill_type="solid")
PURPLE_ACCENT_FILL = PatternFill(bgColor=Color(theme=7, tint=0.4), fill_type="solid")

def apply_formatting(ws):
    """
    Applies conditional formatting and data validation to the worksheet.
    """
    # --- Find column letters ---
    column_letters = {cell.value: cell.column_letter for cell in ws[1]}

    # --- Status Column Formatting ---
    if "Status" in column_letters:
        status_col = column_letters["Status"]
        # Add conditional formatting for each status option
        ws.conditional_formatting.add(f"{status_col}2:{status_col}1048576",
                                    FormulaRule(formula=[f'${status_col}2="{STATUS_OPTIONS[0]}"'], fill=GREEN_FILL)) # COMPLETE
        ws.conditional_formatting.add(f"{status_col}2:{status_col}1048576",
                                    FormulaRule(formula=[f'${status_col}2="{STATUS_OPTIONS[1]}"'], fill=RED_FILL))   # INCOMPLETE
        ws.conditional_formatting.add(f"{status_col}2:{status_col}1048576",
                                    FormulaRule(formula=[f'${status_col}2="{STATUS_OPTIONS[2]}"'], fill=YELLOW_FILL)) # WIP

        # Add dropdown for status options
        status_string = ",".join(STATUS_OPTIONS)
        dv_status = DataValidation(type="list", formula1=f'"{status_string}"', allow_blank=True)
        ws.add_data_validation(dv_status)
        dv_status.add(f"{status_col}2:{status_col}1048576")

    # --- Other Column Formatting ---
    if "Start Date" in column_letters:
        start_date_col = column_letters["Start Date"]
        ws.conditional_formatting.add(f"{start_date_col}2:{start_date_col}1048576",
                                    FormulaRule(formula=[f'NOT(ISBLANK(${start_date_col}2))'], fill=RED_ACCENT_FILL))
    if "End Date" in column_letters:
        end_date_col = column_letters["End Date"]
        ws.conditional_formatting.add(f"{end_date_col}2:{end_date_col}1048576",
                                    FormulaRule(formula=[f'NOT(ISBLANK(${end_date_col}2))'], fill=OLIVE_ACCENT_FILL))
    if "Assigned By" in column_letters:
        assigned_by_col = column_letters["Assigned By"]
        ws.conditional_formatting.add(f"{assigned_by_col}2:{assigned_by_col}1048576",
                                    FormulaRule(formula=[f'NOT(ISBLANK(${assigned_by_col}2))'], fill=PURPLE_ACCENT_FILL))

def main():
    """
    Main function to load, format, and save the Excel file.
    """
    try:
        wb = openpyxl.load_workbook(INPUT_FILENAME)
        ws = wb.active
        apply_formatting(ws)
        wb.save(OUTPUT_FILENAME)
        print(f"Successfully applied formatting. Saved as {OUTPUT_FILENAME}")
    except FileNotFoundError:
        print(f"Error: Input file '{INPUT_FILENAME}' not found. Please run the creation script first.")

if __name__ == "__main__":
    main()
