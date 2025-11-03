import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles.colors import Color

# --- Load the existing workbook ---
# This opens the 'activity_tracker.xlsx' file created by the first script.
wb = openpyxl.load_workbook("activity_tracker.xlsx")
ws = wb.active

# --- Define fill colors for conditional formatting ---
# Standard colors for the "Status" column.
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

# Theme colors for highlighting other columns.
# These colors are based on Excel's theme colors and are lightened by 40% (tint=0.4).
# Accent 2 (theme color 5) is used for a light red.
red_accent_fill = PatternFill(bgColor=Color(theme=5, tint=0.4), fill_type="solid")
# Accent 6 (theme color 9) is used for a light olive green.
olive_accent_fill = PatternFill(bgColor=Color(theme=9, tint=0.4), fill_type="solid")
# Accent 4 (theme color 7) is used for a light purple.
purple_accent_fill = PatternFill(bgColor=Color(theme=7, tint=0.4), fill_type="solid")


# --- Find the column letters for the target columns ---
# This loop iterates through the header row to find the column letters.
# This makes the script more robust if the column order changes.
status_col_letter = None
start_date_col_letter = None
end_date_col_letter = None
assigned_by_col_letter = None

for cell in ws[1]:
    if cell.value == "Status":
        status_col_letter = cell.column_letter
    elif cell.value == "Start Date":
        start_date_col_letter = cell.column_letter
    elif cell.value == "End Date":
        end_date_col_letter = cell.column_letter
    elif cell.value == "Assigned By":
        assigned_by_col_letter = cell.column_letter

# --- Apply conditional formatting and data validation for the "Status" column ---
if status_col_letter:
    # These rules apply a background color to the "Status" cells based on their text content.
    ws.conditional_formatting.add(f"{status_col_letter}2:{status_col_letter}1048576",
                                FormulaRule(formula=[f'${status_col_letter}2="COMPLETE"'], fill=green_fill))
    ws.conditional_formatting.add(f"{status_col_letter}2:{status_col_letter}1048576",
                                FormulaRule(formula=[f'${status_col_letter}2="INCOMPLETE"'], fill=red_fill))
    ws.conditional_formatting.add(f"{status_col_letter}2:{status_col_letter}1048576",
                                FormulaRule(formula=[f'${status_col_letter}2="WIP"'], fill=yellow_fill))
    
    # This adds a dropdown list to the "Status" column for easy data entry.
    dv = openpyxl.worksheet.datavalidation.DataValidation(type="list", formula1='"COMPLETE,INCOMPLETE,WIP"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{status_col_letter}2:{status_col_letter}1048576")

# --- Apply conditional formatting for other columns ---
# These rules highlight the cells in their respective columns if they are not empty.
if start_date_col_letter:
    ws.conditional_formatting.add(f"{start_date_col_letter}2:{start_date_col_letter}1048576",
                                FormulaRule(formula=[f'NOT(ISBLANK(${start_date_col_letter}2))'], fill=red_accent_fill))

if end_date_col_letter:
    ws.conditional_formatting.add(f"{end_date_col_letter}2:{end_date_col_letter}1048576",
                                FormulaRule(formula=[f'NOT(ISBLANK(${end_date_col_letter}2))'], fill=olive_accent_fill))

if assigned_by_col_letter:
    ws.conditional_formatting.add(f"{assigned_by_col_letter}2:{assigned_by_col_letter}1048576",
                                FormulaRule(formula=[f'NOT(ISBLANK(${assigned_by_col_letter}2))'], fill=purple_accent_fill))


# --- Save the final workbook ---
# The workbook with all the applied formatting is saved to a new file.
wb.save("activity_tracker_formatted_2.xlsx")

print("Successfully applied conditional formatting. Saved as activity_tracker_formatted_2.xlsx")